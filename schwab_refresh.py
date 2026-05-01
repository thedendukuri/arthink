#!/usr/bin/env python3
"""
schwab_refresh.py — Refresh options_dashboard.xlsx with Schwab Developer API
data (real-time, free with a Schwab brokerage account).

PREREQUISITES:
  1. Active Schwab brokerage account.
  2. App registered + approved at https://developer.schwab.com (1–3 day wait).
     - Add product: "Market Data Production"
     - Callback URL: https://127.0.0.1:8182
  3. Save your App Key + App Secret in a .env file alongside this script (see
     .env.example).

FIRST RUN:
  Opens a browser. You log into Schwab, authorize the app, then copy the URL
  the browser is redirected to (it'll look like
  `https://127.0.0.1:8182/?code=...`) and paste it back into the terminal.
  The script extracts the auth code, exchanges it for tokens, and saves them
  to schwab_tokens.json. After that, refresh runs are non-interactive.

SUBSEQUENT RUNS:
  Reads schwab_tokens.json. Access tokens expire after 30 min — auto-refreshed
  using the refresh token. Refresh tokens expire after 7 days — when that
  happens, the script will prompt you to redo the browser auth step.

USAGE:
  python schwab_refresh.py [path_to_xlsx]

DEPS:
  pip install requests openpyxl pandas numpy python-dotenv --break-system-packages

API DOCS:
  https://developer.schwab.com/products/trader-api--individual/details/specifications/Market%20Data%20Production
"""
from __future__ import annotations
import sys
import os
import json
import time
import math
import base64
import webbrowser
import urllib.parse
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import requests
    import numpy as np
    from openpyxl import load_workbook
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install requests openpyxl pandas numpy python-dotenv --break-system-packages")
    sys.exit(1)

# Optional .env loader — script works without it if you set env vars directly
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# ---------- Schwab API constants ----------
AUTH_BASE = "https://api.schwabapi.com/v1/oauth"
MD_BASE = "https://api.schwabapi.com/marketdata/v1"
DEFAULT_CALLBACK = "https://127.0.0.1:8182"
TOKENS_FILE = Path(__file__).parent / "schwab_tokens.json"


# ---------- OAuth ----------
class SchwabAuth:
    def __init__(self, app_key: str, app_secret: str, callback: str = DEFAULT_CALLBACK):
        self.app_key = app_key
        self.app_secret = app_secret
        self.callback = callback
        self.tokens = self._load_tokens()

    def _load_tokens(self) -> dict:
        if TOKENS_FILE.exists():
            try:
                return json.loads(TOKENS_FILE.read_text())
            except Exception:
                return {}
        return {}

    def _save_tokens(self, data: dict):
        # Augment with computed expiry timestamps
        now = datetime.now(timezone.utc)
        if "expires_in" in data:
            data["access_token_expires_at"] = (now + timedelta(seconds=data["expires_in"] - 60)).isoformat()
        # Refresh tokens nominally last 7 days
        data["refresh_token_expires_at"] = (now + timedelta(days=6, hours=20)).isoformat()
        TOKENS_FILE.write_text(json.dumps(data, indent=2))
        self.tokens = data

    def _basic_auth_header(self) -> dict:
        creds = f"{self.app_key}:{self.app_secret}".encode()
        return {"Authorization": "Basic " + base64.b64encode(creds).decode()}

    def _interactive_authorize(self):
        """Open browser, prompt user to paste redirect URL, exchange code for tokens."""
        auth_url = (
            f"{AUTH_BASE}/authorize?client_id={self.app_key}"
            f"&redirect_uri={urllib.parse.quote(self.callback, safe='')}"
        )
        print("\n" + "=" * 70)
        print("Opening Schwab authorization page in your browser...")
        print("If it doesn't open, paste this URL manually:")
        print(f"  {auth_url}")
        print("=" * 70)
        try:
            webbrowser.open(auth_url)
        except Exception:
            pass
        print("\nAfter you log in and approve, your browser will redirect to a URL")
        print(f"starting with {self.callback} (it may show a 'site can't be reached'")
        print("error — that's expected). Copy the FULL URL from the address bar")
        print("and paste it here:\n")
        redirected = input("Paste redirect URL: ").strip()

        # Extract `code` query param
        parsed = urllib.parse.urlparse(redirected)
        params = urllib.parse.parse_qs(parsed.query)
        if "code" not in params:
            raise RuntimeError("No 'code' parameter in pasted URL. Try again.")
        code = params["code"][0]

        # Exchange code for tokens
        resp = requests.post(
            f"{AUTH_BASE}/token",
            headers={**self._basic_auth_header(),
                     "Content-Type": "application/x-www-form-urlencoded"},
            data={
                "grant_type": "authorization_code",
                "code": code,
                "redirect_uri": self.callback,
            },
            timeout=30,
        )
        if resp.status_code != 200:
            raise RuntimeError(f"Token exchange failed: {resp.status_code} {resp.text}")
        self._save_tokens(resp.json())
        print("Authorization successful — tokens saved.\n")

    def _refresh_access_token(self):
        if "refresh_token" not in self.tokens:
            raise RuntimeError("No refresh token. Re-run authorization.")
        resp = requests.post(
            f"{AUTH_BASE}/token",
            headers={**self._basic_auth_header(),
                     "Content-Type": "application/x-www-form-urlencoded"},
            data={
                "grant_type": "refresh_token",
                "refresh_token": self.tokens["refresh_token"],
            },
            timeout=30,
        )
        if resp.status_code != 200:
            print(f"Refresh failed ({resp.status_code}): refresh token may be expired.")
            print("Falling back to interactive auth...")
            self._interactive_authorize()
            return
        # Refresh response keeps the same refresh_token
        new = resp.json()
        if "refresh_token" not in new and "refresh_token" in self.tokens:
            new["refresh_token"] = self.tokens["refresh_token"]
        self._save_tokens(new)

    def access_token(self) -> str:
        if not self.tokens:
            self._interactive_authorize()
        # Check expiry
        exp_str = self.tokens.get("access_token_expires_at")
        if exp_str:
            try:
                exp = datetime.fromisoformat(exp_str)
                if exp <= datetime.now(timezone.utc):
                    self._refresh_access_token()
            except Exception:
                self._refresh_access_token()
        else:
            self._refresh_access_token()
        return self.tokens["access_token"]

    def auth_header(self) -> dict:
        return {"Authorization": f"Bearer {self.access_token()}"}


# ---------- Schwab market data calls ----------
def schwab_quotes(auth: SchwabAuth, symbols: list[str]) -> dict:
    """GET /marketdata/v1/quotes — batch quotes."""
    if not symbols:
        return {}
    resp = requests.get(
        f"{MD_BASE}/quotes",
        headers=auth.auth_header(),
        params={"symbols": ",".join(symbols), "fields": "quote,reference"},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def schwab_chain(auth: SchwabAuth, symbol: str, expiration: str | None = None) -> dict:
    """GET /marketdata/v1/chains — full options chain for a symbol."""
    params = {"symbol": symbol, "contractType": "ALL", "includeUnderlyingQuote": True}
    if expiration:
        # Schwab accepts a single date for fromDate=toDate
        params["fromDate"] = expiration
        params["toDate"] = expiration
    else:
        # Default: next 60 days of expirations; we'll pick nearest
        today = datetime.now().date()
        params["fromDate"] = today.isoformat()
        params["toDate"] = (today + timedelta(days=90)).isoformat()
    resp = requests.get(
        f"{MD_BASE}/chains",
        headers=auth.auth_header(),
        params=params,
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def schwab_price_history(auth: SchwabAuth, symbol: str, days: int = 35) -> list:
    """GET /marketdata/v1/{symbol}/pricehistory — daily candles for HV calc."""
    end_ms = int(time.time() * 1000)
    start_ms = end_ms - days * 24 * 60 * 60 * 1000
    resp = requests.get(
        f"{MD_BASE}/pricehistory",
        headers=auth.auth_header(),
        params={
            "symbol": symbol,
            "periodType": "month",
            "frequencyType": "daily",
            "frequency": 1,
            "startDate": start_ms,
            "endDate": end_ms,
        },
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json().get("candles", [])


def hv30(auth: SchwabAuth, symbol: str) -> float | None:
    try:
        candles = schwab_price_history(auth, symbol, 35)
        closes = [c["close"] for c in candles if c.get("close")]
        if len(closes) < 5:
            return None
        rets = np.diff(np.log(closes))
        return float(rets.std() * math.sqrt(252))
    except Exception:
        return None


# ---------- Position parser (shared with yfinance script) ----------
def parse_position_option(s):
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    if " " in s:
        parts = s.split()
        if len(parts) >= 4:
            try:
                underlying = parts[0].upper()
                exp = parts[1]
                kind = parts[2].lower()
                strike = float(parts[3])
                if "c" in kind:
                    kind = "Call"
                elif "p" in kind:
                    kind = "Put"
                else:
                    return None
                datetime.strptime(exp, "%Y-%m-%d")
                return underlying, exp, kind, strike
            except Exception:
                return None
    import re
    m = re.match(r"^([A-Z\.]{1,6})(\d{6})([CP])(\d{8})$", s)
    if not m:
        return None
    underlying, ymd, cp, strike_str = m.groups()
    yy, mm, dd = ymd[:2], ymd[2:4], ymd[4:6]
    year = 2000 + int(yy)
    return underlying, f"{year:04d}-{mm}-{dd}", "Call" if cp == "C" else "Put", int(strike_str) / 1000.0


def schwab_option_symbol(underlying: str, exp_yyyy_mm_dd: str, kind: str, strike: float) -> str:
    """Build Schwab's OSI option symbol format: 'AAPL  250620C00200000' (note 2 spaces after root)."""
    # Schwab uses fixed-width: 6-char root padded, YYMMDD, C/P, 8-digit strike (× 1000)
    root = underlying.ljust(6)
    yy, mm, dd = exp_yyyy_mm_dd.split("-")
    yy = yy[-2:]
    cp = "C" if kind.lower().startswith("c") else "P"
    k = f"{int(round(strike * 1000)):08d}"
    return f"{root}{yy}{mm}{dd}{cp}{k}"


# ---------- Refresh ----------
def refresh(xlsx_path: Path, auth: SchwabAuth) -> dict:
    api_calls = 0
    started = datetime.now(timezone.utc)

    wb = load_workbook(xlsx_path)
    config = wb["Config"]
    watchlist = wb["Watchlist"]

    # ---- Watchlist (batched) ----
    tickers = []
    for row in range(4, watchlist.max_row + 1):
        v = watchlist.cell(row=row, column=1).value
        if v and str(v).strip():
            tickers.append((row, str(v).strip().upper()))

    print(f"Watchlist: {len(tickers)} tickers — fetching batch quotes...")
    syms = [s for _, s in tickers]
    quotes_data = schwab_quotes(auth, syms)
    api_calls += 1

    refreshed = 0
    for row, sym in tickers:
        q = quotes_data.get(sym, {})
        quote = q.get("quote", {})
        ref = q.get("reference", {})
        last = quote.get("lastPrice")
        prev = quote.get("closePrice") or quote.get("regularMarketPreviousClose")
        bid = quote.get("bidPrice")
        ask = quote.get("askPrice")
        high = quote.get("highPrice")
        low = quote.get("lowPrice")
        vol = quote.get("totalVolume")
        change = (last - prev) if (last is not None and prev) else None
        change_pct = (change / prev) if (change is not None and prev) else None

        watchlist.cell(row=row, column=2,  value=last)
        watchlist.cell(row=row, column=3,  value=change)
        watchlist.cell(row=row, column=4,  value=change_pct)
        watchlist.cell(row=row, column=5,  value=bid)
        watchlist.cell(row=row, column=6,  value=ask)
        watchlist.cell(row=row, column=7,  value=high)
        watchlist.cell(row=row, column=8,  value=low)
        watchlist.cell(row=row, column=9,  value=vol)
        watchlist.cell(row=row, column=10, value=prev)
        # 30D HV — separate call per ticker; can be slow, skip if HV slot already fresh today
        watchlist.cell(row=row, column=11, value=hv30(auth, sym))
        api_calls += 1
        refreshed += 1
        time.sleep(0.05)
        if refreshed % 10 == 0:
            print(f"  ...{refreshed}/{len(tickers)}")

    # ---- Options Chain ----
    if "OptionsChain" in wb.sheetnames:
        oc = wb["OptionsChain"]
        chain_ticker = oc.cell(row=3, column=2).value
        chain_exp = oc.cell(row=3, column=4).value
        if chain_ticker:
            chain_ticker = str(chain_ticker).strip().upper()
            try:
                data = schwab_chain(auth, chain_ticker,
                                    str(chain_exp).strip() if chain_exp else None)
                api_calls += 1
                # Schwab response: callExpDateMap / putExpDateMap
                # Each map: { "YYYY-MM-DD:DTE": { "strike": [contract, ...], ... } }
                spot = data.get("underlyingPrice") or data.get("underlying", {}).get("last")

                call_map = data.get("callExpDateMap", {})
                put_map = data.get("putExpDateMap", {})

                # Pick expiration: user-specified if present, otherwise nearest in callExp
                if chain_exp and isinstance(chain_exp, str):
                    target_exp = chain_exp.strip()
                    exp_keys = [k for k in call_map.keys() if k.startswith(target_exp)]
                    if not exp_keys:
                        exp_keys = sorted(call_map.keys())[:1]
                else:
                    exp_keys = sorted(call_map.keys())[:1]
                if not exp_keys:
                    print(f"OptionsChain: no expirations returned for {chain_ticker}")
                else:
                    exp_key = exp_keys[0]
                    used_exp = exp_key.split(":")[0]
                    oc.cell(row=3, column=4, value=used_exp)
                    oc.cell(row=3, column=6, value=spot)
                    oc.cell(row=3, column=8, value=started.strftime("%Y-%m-%d %H:%M UTC"))

                    # Find matching put expiration key (Schwab uses same date but DTE could differ trivially)
                    put_keys = [k for k in put_map.keys() if k.startswith(used_exp)]
                    put_exp_key = put_keys[0] if put_keys else None

                    call_strikes = call_map.get(exp_key, {})
                    put_strikes = put_map.get(put_exp_key, {}) if put_exp_key else {}

                    all_strikes = sorted(set(float(k) for k in call_strikes.keys()) |
                                         set(float(k) for k in put_strikes.keys()))
                    if spot:
                        all_strikes = sorted(all_strikes, key=lambda k: abs(k - spot))[:60]
                        all_strikes = sorted(all_strikes)

                    HDR = 5
                    for r in range(HDR + 1, HDR + 61):
                        for c in range(1, 14):
                            oc.cell(row=r, column=c, value=None)

                    for i, k in enumerate(all_strikes[:60]):
                        rr = HDR + 1 + i
                        cs = call_strikes.get(str(k)) or call_strikes.get(f"{k:.1f}") or call_strikes.get(f"{k:.2f}")
                        ps = put_strikes.get(str(k)) or put_strikes.get(f"{k:.1f}") or put_strikes.get(f"{k:.2f}")
                        c = cs[0] if cs else None
                        p = ps[0] if ps else None
                        if c:
                            oc.cell(row=rr, column=1, value=c.get("last"))
                            oc.cell(row=rr, column=2, value=c.get("bid"))
                            oc.cell(row=rr, column=3, value=c.get("ask"))
                            oc.cell(row=rr, column=4, value=c.get("totalVolume"))
                            oc.cell(row=rr, column=5, value=c.get("openInterest"))
                            iv = c.get("volatility")
                            oc.cell(row=rr, column=6, value=iv / 100 if iv else None)
                        oc.cell(row=rr, column=7, value=k)
                        if p:
                            oc.cell(row=rr, column=8,  value=p.get("last"))
                            oc.cell(row=rr, column=9,  value=p.get("bid"))
                            oc.cell(row=rr, column=10, value=p.get("ask"))
                            oc.cell(row=rr, column=11, value=p.get("totalVolume"))
                            oc.cell(row=rr, column=12, value=p.get("openInterest"))
                            iv = p.get("volatility")
                            oc.cell(row=rr, column=13, value=iv / 100 if iv else None)
                    print(f"OptionsChain: {chain_ticker} {used_exp} — {len(all_strikes[:60])} strikes")
            except Exception as e:
                print(f"OptionsChain ERROR for {chain_ticker}: {e}")

    # ---- Positions (option marks via batch quote on OSI symbols) ----
    if "Positions" in wb.sheetnames:
        positions = wb["Positions"]
        opt_symbols = []
        rows_to_update = []
        for row in range(4, 19):
            cell_val = positions.cell(row=row, column=3).value
            parsed = parse_position_option(cell_val) if cell_val else None
            if parsed:
                u, exp, kind, strike = parsed
                osi = schwab_option_symbol(u, exp, kind, strike)
                opt_symbols.append(osi)
                rows_to_update.append((row, osi))

        if opt_symbols:
            try:
                opt_quotes = schwab_quotes(auth, opt_symbols)
                api_calls += 1
                for row, osi in rows_to_update:
                    q = opt_quotes.get(osi, {})
                    quote = q.get("quote", {})
                    bid = quote.get("bidPrice")
                    ask = quote.get("askPrice")
                    last = quote.get("lastPrice")
                    if bid and ask and bid > 0 and ask > 0:
                        mark = (bid + ask) / 2
                    else:
                        mark = last
                    if mark is not None:
                        positions.cell(row=row, column=9, value=mark)
            except Exception as e:
                print(f"Positions option-quote ERROR: {e}")

    # ---- Config status ----
    config.cell(row=12, column=2, value=started.strftime("%Y-%m-%d %H:%M:%S"))
    config.cell(row=13, column=2, value="OK (Schwab)")
    config.cell(row=14, column=2, value=refreshed)
    config.cell(row=15, column=2, value=api_calls)

    wb.save(xlsx_path)
    return {"refreshed": refreshed, "api_calls": api_calls}


def main():
    app_key = os.environ.get("SCHWAB_APP_KEY")
    app_secret = os.environ.get("SCHWAB_APP_SECRET")
    callback = os.environ.get("SCHWAB_CALLBACK", DEFAULT_CALLBACK)

    if not app_key or not app_secret:
        print("Missing SCHWAB_APP_KEY / SCHWAB_APP_SECRET in environment.")
        print("Create a .env file (see .env.example) or `export` them in your shell.")
        sys.exit(1)

    auth = SchwabAuth(app_key, app_secret, callback)

    xlsx = Path(sys.argv[1] if len(sys.argv) > 1 else "options_dashboard.xlsx").resolve()
    if not xlsx.exists():
        print(f"File not found: {xlsx}")
        sys.exit(1)

    print(f"Refreshing {xlsx} via Schwab API...")
    try:
        result = refresh(xlsx, auth)
        print(f"Done — {result['refreshed']} tickers, {result['api_calls']} API calls.")
    except Exception as e:
        import traceback
        traceback.print_exc()
        try:
            wb = load_workbook(xlsx)
            cfg = wb["Config"]
            cfg.cell(row=12, column=2,
                     value=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"))
            cfg.cell(row=13, column=2, value=f"ERROR: {e}")
            wb.save(xlsx)
        except Exception:
            pass
        sys.exit(1)


if __name__ == "__main__":
    main()
