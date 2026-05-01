#!/usr/bin/env python3
"""
yfinance_refresh.py — Refresh options_dashboard.xlsx with FREE data.

Data source: yfinance (https://github.com/ranaroussi/yfinance) — unofficial Yahoo
Finance wrapper. No API key required. Quotes are typically delayed 15–20 minutes
on Yahoo's free feed.

Populates:
  • Watchlist sheet — last/change/bid/ask/vol/prev close/30D HV/IV rank
  • OptionsChain sheet — chain for the ticker + expiration set in the sheet's
    selector cells (B3 ticker, D3 expiration). Leave D3 blank for nearest expiry.
  • Positions sheet — Mark column (col I) for any row where col C has an option
    ticker in OCC format (e.g., "AAPL250620C00200000") OR a "TICKER YYYY-MM-DD CALL/PUT STRIKE" string.
  • Config sheet — Last refresh timestamp, status, ticker count, API call count.

Usage:
    python yfinance_refresh.py [path_to_xlsx]

If no path given, defaults to ./options_dashboard.xlsx in the current directory.

Cron / Task Scheduler:
    macOS/Linux cron — refresh every 15 min during market hours:
        */15 9-16 * * 1-5  cd /path/to/folder && python yfinance_refresh.py
    Windows Task Scheduler — point to python.exe with this script as argument.

Limitations:
    • yfinance scrapes Yahoo; structure can change. If quotes return None, try
      `pip install --upgrade yfinance`.
    • No real-time level — for sub-minute trading decisions you need a paid feed
      (Polygon, Tradier funded, Schwab brokerage API).
    • Yahoo rate-limits aggressive scrapers. Keep watchlist under ~50 tickers
      for reliability.

Install:
    pip install yfinance openpyxl pandas numpy --break-system-packages
"""
from __future__ import annotations
import sys
import time
import math
import traceback
from datetime import datetime, timezone
from pathlib import Path

# Lazy-import heavy deps only when running as a script — keeps parse_position_option
# importable for testing without yfinance installed.
def _require_deps():
    global yf, pd, np, load_workbook
    try:
        import yfinance as yf  # noqa: F401
        import pandas as pd  # noqa: F401
        import numpy as np  # noqa: F401
        from openpyxl import load_workbook  # noqa: F401
    except ImportError as e:
        print(f"Missing dependency: {e}")
        print("Run: pip install yfinance openpyxl pandas numpy --break-system-packages")
        sys.exit(1)
    globals()["yf"] = yf
    globals()["pd"] = pd
    globals()["np"] = np
    globals()["load_workbook"] = load_workbook


# ---------- Helpers ----------
def _safe(d, *keys, default=None):
    """Walk a nested dict, return default if any key missing or value None."""
    cur = d
    for k in keys:
        if cur is None:
            return default
        if isinstance(cur, dict):
            cur = cur.get(k)
        else:
            return default
    return cur if cur is not None else default


def hist_volatility(ticker: yf.Ticker, days: int = 30) -> float | None:
    """Annualized realized volatility from daily close returns."""
    try:
        h = ticker.history(period=f"{days+5}d", auto_adjust=False)
        if h is None or len(h) < 5:
            return None
        rets = np.log(h["Close"] / h["Close"].shift(1)).dropna()
        if len(rets) < 5:
            return None
        return float(rets.std() * math.sqrt(252))
    except Exception:
        return None


def iv_rank(ticker: yf.Ticker) -> float | None:
    """
    Crude IV rank proxy: where the nearest-expiry ATM call IV sits within the
    range of (HV30, HV30*2). Real IV rank needs 1y of IV history which yfinance
    doesn't provide. This is a placeholder — treat as directional only.
    """
    try:
        expirations = ticker.options
        if not expirations:
            return None
        chain = ticker.option_chain(expirations[0])
        calls = chain.calls
        if calls is None or calls.empty:
            return None
        spot = ticker.fast_info.get("last_price") or ticker.fast_info.get("lastPrice")
        if not spot:
            return None
        atm = calls.iloc[(calls["strike"] - spot).abs().argsort()[:1]]
        iv = float(atm["impliedVolatility"].iloc[0])
        hv = hist_volatility(ticker, 30)
        if not hv or hv <= 0:
            return iv  # fall back to raw IV
        # Map iv into [hv, 2*hv] → [0, 1]
        rank = (iv - hv) / (hv) if hv > 0 else None
        if rank is None:
            return iv
        return max(0.0, min(1.0, rank))
    except Exception:
        return None


def _fi_get(fi, key):
    """FastInfo can be a dict, a FastInfo object, or None across yfinance versions."""
    if fi is None:
        return None
    # Dict-like with .get()
    if hasattr(fi, "get"):
        try:
            v = fi.get(key)
            if v is not None:
                return v
        except Exception:
            pass
    # Attribute access
    try:
        v = getattr(fi, key, None)
        if v is not None:
            return v
    except Exception:
        pass
    # Bracket access
    try:
        return fi[key]
    except Exception:
        return None


def fetch_quote(symbol: str) -> dict:
    """Pull a single ticker's snapshot. History-first; fast_info as supplement."""
    t = yf.Ticker(symbol)

    # PRIMARY: pull last 5 trading days via history. Works during market hours,
    # pre/post-market, and weekends. fast_info is unreliable across yfinance versions.
    last = prev_close = day_high = day_low = volume = None
    try:
        hist = t.history(period="5d", auto_adjust=False)
        if hist is not None and not hist.empty:
            last_row = hist.iloc[-1]
            last = float(last_row["Close"]) if pd.notna(last_row["Close"]) else None
            day_high = float(last_row["High"]) if pd.notna(last_row["High"]) else None
            day_low = float(last_row["Low"]) if pd.notna(last_row["Low"]) else None
            volume = int(last_row["Volume"]) if pd.notna(last_row["Volume"]) else None
            if len(hist) >= 2:
                prev_close = float(hist.iloc[-2]["Close"])
    except Exception as e:
        print(f"  {symbol}: history error: {e}")

    # SUPPLEMENT: try fast_info for live bid/ask + intraday last (only during market hrs)
    bid = ask = None
    try:
        fi = t.fast_info
        bid = _fi_get(fi, "bid")
        ask = _fi_get(fi, "ask")
        fi_last = _fi_get(fi, "last_price") or _fi_get(fi, "lastPrice")
        if fi_last is not None:
            try:
                last = float(fi_last)
            except (TypeError, ValueError):
                pass
        fi_prev = _fi_get(fi, "previous_close") or _fi_get(fi, "previousClose")
        if fi_prev is not None and prev_close is None:
            try:
                prev_close = float(fi_prev)
            except (TypeError, ValueError):
                pass
    except Exception:
        pass

    change = (last - prev_close) if (last is not None and prev_close) else None
    change_pct = (change / prev_close) if (change is not None and prev_close) else None

    return {
        "symbol": symbol,
        "last": last,
        "change": change,
        "change_pct": change_pct,
        "bid": bid,
        "ask": ask,
        "day_high": day_high,
        "day_low": day_low,
        "volume": volume,
        "prev_close": prev_close,
        "hv30": hist_volatility(t, 30),
        "iv_rank": iv_rank(t),
        "_ticker": t,
    }


def fetch_chain(symbol: str, expiration: str | None = None):
    """Returns (chain_df_calls, chain_df_puts, chosen_expiry, underlying_price)."""
    t = yf.Ticker(symbol)
    expirations = list(t.options or [])
    if not expirations:
        return None, None, None, None

    if expiration and expiration in expirations:
        exp = expiration
    else:
        exp = expirations[0]  # nearest

    chain = t.option_chain(exp)
    spot = t.fast_info.get("last_price") or t.fast_info.get("lastPrice")
    return chain.calls, chain.puts, exp, spot


def parse_position_option(s: str):
    """
    Parse a position 'Option Ticker' cell into (underlying, expiry_yyyymmdd, type, strike).

    Accepts two formats:
      1) OCC standard: 'AAPL250620C00200000'  → AAPL, 2025-06-20, Call, 200.00
      2) Loose:        'AAPL 2025-06-20 CALL 200'
    """
    if not s or not isinstance(s, str):
        return None
    s = s.strip()

    # Loose format with spaces
    if " " in s:
        parts = s.split()
        if len(parts) >= 4:
            try:
                underlying = parts[0].upper()
                exp = parts[1]
                kind = parts[2].lower()
                strike = float(parts[3])
                if "c" in kind:
                    kind = "Call"
                elif "p" in kind:
                    kind = "Put"
                else:
                    return None
                # Normalize date
                datetime.strptime(exp, "%Y-%m-%d")
                return underlying, exp, kind, strike
            except Exception:
                return None

    # OCC standard
    # Underlying chars + YYMMDD + C/P + 8-digit strike (×1000)
    import re
    m = re.match(r"^([A-Z\.]{1,6})(\d{6})([CP])(\d{8})$", s)
    if not m:
        return None
    underlying, ymd, cp, strike_str = m.groups()
    yy, mm, dd = ymd[:2], ymd[2:4], ymd[4:6]
    year = 2000 + int(yy)
    exp = f"{year:04d}-{mm}-{dd}"
    strike = int(strike_str) / 1000.0
    return underlying, exp, "Call" if cp == "C" else "Put", strike


def lookup_option_mark(underlying: str, exp: str, kind: str, strike: float) -> float | None:
    try:
        t = yf.Ticker(underlying)
        if exp not in (t.options or []):
            return None
        chain = t.option_chain(exp)
        df = chain.calls if kind == "Call" else chain.puts
        if df is None or df.empty:
            return None
        # Match nearest strike
        idx = (df["strike"] - strike).abs().idxmin()
        row = df.loc[idx]
        bid = row.get("bid")
        ask = row.get("ask")
        last = row.get("lastPrice")
        if bid and ask and bid > 0 and ask > 0:
            return float((bid + ask) / 2)
        return float(last) if last else None
    except Exception:
        return None


# ---------- Main refresh ----------
def refresh(xlsx_path: Path) -> dict:
    api_calls = 0
    started = datetime.now(timezone.utc)

    wb = load_workbook(xlsx_path)
    if "Config" not in wb.sheetnames or "Watchlist" not in wb.sheetnames:
        raise RuntimeError("Workbook is missing required sheets (Config/Watchlist).")

    config = wb["Config"]
    watchlist = wb["Watchlist"]

    # ---- Watchlist refresh ----
    tickers = []
    for row in range(4, watchlist.max_row + 1):
        v = watchlist.cell(row=row, column=1).value
        if v and str(v).strip():
            tickers.append((row, str(v).strip().upper()))

    print(f"Watchlist: {len(tickers)} tickers")
    refreshed = 0
    for row, sym in tickers:
        try:
            q = fetch_quote(sym)
            api_calls += 3  # fast_info + history + option_chain calls inside
            watchlist.cell(row=row, column=2,  value=q["last"])
            watchlist.cell(row=row, column=3,  value=q["change"])
            watchlist.cell(row=row, column=4,  value=q["change_pct"])
            watchlist.cell(row=row, column=5,  value=q["bid"])
            watchlist.cell(row=row, column=6,  value=q["ask"])
            watchlist.cell(row=row, column=7,  value=q["day_high"])
            watchlist.cell(row=row, column=8,  value=q["day_low"])
            watchlist.cell(row=row, column=9,  value=q["volume"])
            watchlist.cell(row=row, column=10, value=q["prev_close"])
            watchlist.cell(row=row, column=11, value=q["hv30"])
            watchlist.cell(row=row, column=12, value=q["iv_rank"])
            refreshed += 1
            print(f"  {sym}: last={q['last']} chg%={q['change_pct']}")
        except Exception as e:
            print(f"  {sym}: ERROR {e}")
        time.sleep(0.2)  # gentle rate limit

    # ---- Options Chain refresh ----
    if "OptionsChain" in wb.sheetnames:
        oc = wb["OptionsChain"]
        chain_ticker = oc.cell(row=3, column=2).value
        chain_exp = oc.cell(row=3, column=4).value
        if chain_ticker:
            chain_ticker = str(chain_ticker).strip().upper()
            try:
                calls, puts, used_exp, spot = fetch_chain(chain_ticker, str(chain_exp).strip() if chain_exp else None)
                api_calls += 2
                if calls is not None and puts is not None:
                    oc.cell(row=3, column=4, value=used_exp)
                    oc.cell(row=3, column=6, value=spot)
                    oc.cell(row=3, column=8, value=started.strftime("%Y-%m-%d %H:%M UTC"))

                    # Build strike-aligned merged view
                    calls_idx = calls.set_index("strike")
                    puts_idx = puts.set_index("strike")
                    all_strikes = sorted(set(calls_idx.index) | set(puts_idx.index))

                    # Limit to ~30 strikes around spot to fit in 60-row reservation
                    if spot:
                        all_strikes = sorted(all_strikes, key=lambda k: abs(k - spot))[:60]
                        all_strikes = sorted(all_strikes)

                    HDR = 5
                    # Clear existing rows
                    for r in range(HDR + 1, HDR + 61):
                        for c in range(1, 14):
                            oc.cell(row=r, column=c, value=None)

                    for i, k in enumerate(all_strikes[:60]):
                        rr = HDR + 1 + i
                        c = calls_idx.loc[k] if k in calls_idx.index else None
                        p = puts_idx.loc[k] if k in puts_idx.index else None
                        if c is not None and isinstance(c, pd.DataFrame):
                            c = c.iloc[0]
                        if p is not None and isinstance(p, pd.DataFrame):
                            p = p.iloc[0]
                        if c is not None:
                            oc.cell(row=rr, column=1, value=float(c.get("lastPrice")) if pd.notna(c.get("lastPrice")) else None)
                            oc.cell(row=rr, column=2, value=float(c.get("bid")) if pd.notna(c.get("bid")) else None)
                            oc.cell(row=rr, column=3, value=float(c.get("ask")) if pd.notna(c.get("ask")) else None)
                            oc.cell(row=rr, column=4, value=int(c.get("volume")) if pd.notna(c.get("volume")) else None)
                            oc.cell(row=rr, column=5, value=int(c.get("openInterest")) if pd.notna(c.get("openInterest")) else None)
                            oc.cell(row=rr, column=6, value=float(c.get("impliedVolatility")) if pd.notna(c.get("impliedVolatility")) else None)
                        oc.cell(row=rr, column=7, value=float(k))
                        if p is not None:
                            oc.cell(row=rr, column=8,  value=float(p.get("lastPrice")) if pd.notna(p.get("lastPrice")) else None)
                            oc.cell(row=rr, column=9,  value=float(p.get("bid")) if pd.notna(p.get("bid")) else None)
                            oc.cell(row=rr, column=10, value=float(p.get("ask")) if pd.notna(p.get("ask")) else None)
                            oc.cell(row=rr, column=11, value=int(p.get("volume")) if pd.notna(p.get("volume")) else None)
                            oc.cell(row=rr, column=12, value=int(p.get("openInterest")) if pd.notna(p.get("openInterest")) else None)
                            oc.cell(row=rr, column=13, value=float(p.get("impliedVolatility")) if pd.notna(p.get("impliedVolatility")) else None)
                    print(f"OptionsChain: {chain_ticker} {used_exp} — {len(all_strikes[:60])} strikes")
                else:
                    print(f"OptionsChain: no chain returned for {chain_ticker}")
            except Exception as e:
                print(f"OptionsChain ERROR for {chain_ticker}: {e}")

    # ---- Positions mark refresh ----
    if "Positions" in wb.sheetnames:
        positions = wb["Positions"]
        for row in range(4, 19):
            cell_val = positions.cell(row=row, column=3).value  # column C: Option Ticker
            parsed = parse_position_option(cell_val) if cell_val else None
            if parsed:
                u, exp, kind, strike = parsed
                try:
                    mark = lookup_option_mark(u, exp, kind, strike)
                    api_calls += 2
                    if mark is not None:
                        positions.cell(row=row, column=9, value=mark)
                except Exception as e:
                    print(f"Positions row {row} ({cell_val}): {e}")

    # ---- Config status ----
    config.cell(row=12, column=2, value=started.strftime("%Y-%m-%d %H:%M:%S"))  # Last refresh (UTC)
    config.cell(row=13, column=2, value="OK")
    config.cell(row=14, column=2, value=refreshed)
    config.cell(row=15, column=2, value=api_calls)

    wb.save(xlsx_path)
    return {"refreshed": refreshed, "api_calls": api_calls}


def main():
    _require_deps()
    xlsx = Path(sys.argv[1] if len(sys.argv) > 1 else "options_dashboard.xlsx").resolve()
    if not xlsx.exists():
        print(f"File not found: {xlsx}")
        sys.exit(1)
    print(f"Refreshing {xlsx}")
    try:
        result = refresh(xlsx)
        print(f"Done — {result['refreshed']} tickers refreshed, ~{result['api_calls']} API calls.")
    except Exception as e:
        traceback.print_exc()
        # Mark failure in Config sheet
        try:
            wb = load_workbook(xlsx)
            cfg = wb["Config"]
            cfg.cell(row=12, column=2, value=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"))
            cfg.cell(row=13, column=2, value=f"ERROR: {e}")
            wb.save(xlsx)
        except Exception:
            pass
        sys.exit(1)


if __name__ == "__main__":
    main()
