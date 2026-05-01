"""
Build options_dashboard.xlsx — a multi-sheet options trading dashboard
that pairs with polygon_refresh.py for live data refresh.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName

# ---------- Style helpers ----------
FONT_NAME = "Arial"

HEADER_FILL = PatternFill("solid", start_color="1F3864")
SUBHEADER_FILL = PatternFill("solid", start_color="2E75B6")
INPUT_FILL = PatternFill("solid", start_color="FFF2CC")    # yellow = key input
WARN_FILL = PatternFill("solid", start_color="FCE4D6")     # light red
GOOD_FILL = PatternFill("solid", start_color="E2EFDA")     # light green
BAND_FILL = PatternFill("solid", start_color="F2F2F2")     # zebra

HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F3864")
BODY_FONT = Font(name=FONT_NAME, size=10)
INPUT_FONT = Font(name=FONT_NAME, size=10, color="0000FF", bold=True)  # blue = input
FORMULA_FONT = Font(name=FONT_NAME, size=10, color="000000")           # black = formula
LINK_FONT = Font(name=FONT_NAME, size=10, color="008000")              # green = cross-sheet
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="595959")

THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header(cell, fill=HEADER_FILL, font=HEADER_FONT):
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER
    cell.border = BOX


def write_title(ws, row, col, text, span=8):
    ws.cell(row=row, column=col, value=text).font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)


# ---------- Build workbook ----------
wb = Workbook()

# =================================================================
# SHEET 1: README
# =================================================================
ws = wb.active
ws.title = "README"
set_col_widths(ws, [3, 100])

write_title(ws, 1, 2, "Options Trading Dashboard — Read Me First", span=1)
ws.cell(row=2, column=2,
        value="A Polygon.io-powered options dashboard for tracking S&P 500 names, "
              "modeling strategies, and managing risk on a small account.").font = NOTE_FONT

readme_rows = [
    ("",  ""),
    ("RISK WARNING",
     "Options trading carries substantial risk of loss, including total loss of "
     "principal, and can result in losses exceeding deposited capital when using "
     "uncovered/short strategies. Past performance does not predict future results. "
     "Targeting 10x+ returns over short periods (e.g., $1k → $10k+ in 2 months) is "
     "statistically improbable; historically, the strategies capable of such returns "
     "(0DTE far-OTM lottos, naked positions) have base rates that wipe out the "
     "overwhelming majority of accounts. This tool is informational. It is not "
     "investment advice."),
    ("",  ""),
    ("SETUP — 5 STEPS",
     ""),
    ("1. Get a Polygon API key",
     "Sign up at https://polygon.io. Free tier gives end-of-day stock data with "
     "delays. To get real-time options data you need at minimum the Options Starter "
     "plan (~$29–99/mo at time of writing — verify current pricing). Without a paid "
     "options tier, the chain refresh will return 15-min delayed or limited data."),
    ("2. Install Python dependencies",
     "Run: pip install openpyxl requests pandas polygon-api-client --break-system-packages"),
    ("3. Paste your API key into the Config sheet",
     "Go to the Config sheet, cell B3, and paste your Polygon API key into the yellow cell. "
     "Set your account size in B4. Set your max-risk-per-trade % in B5."),
    ("4. Edit your watchlist",
     "Go to Watchlist sheet. Tickers in column A drive what the refresh script pulls. "
     "Pre-filled with major indices and the most-liquid S&P 500 names. "
     "You can add/remove rows. Saving the file is required before refresh."),
    ("5. Refresh data",
     "From terminal in this folder run: python polygon_refresh.py — it reads your "
     "config, pulls quotes + options chains for the watchlist, writes into this "
     "workbook, and saves. Re-run any time. Set up cron/Task Scheduler for "
     "automatic refresh (see polygon_refresh.py header for instructions)."),
    ("",  ""),
    ("SHEET MAP",
     ""),
    ("• Config",
     "API key, account size, risk parameters, refresh log."),
    ("• Watchlist",
     "Live quotes (last, change, %, bid/ask, volume, IV rank) for the tickers you track. "
     "Refreshed by polygon_refresh.py."),
    ("• OptionsChain",
     "Filtered options chain for one selected ticker + expiration. Change the "
     "ticker/expiry in the yellow cells, then re-run refresh to pull that chain."),
    ("• StrategyCalc",
     "Pick a strategy template (Long Call, Long Put, Bull Call Spread, Bear Put Spread, "
     "Iron Condor, Cash-Secured Put, Covered Call, Straddle). Input legs and underlying "
     "price. Sheet computes max profit, max loss, breakevens, and risk/reward."),
    ("• PositionSizer",
     "Given account size and max-risk-per-trade %, computes max contracts you can buy "
     "for any given option price. Critical for small accounts: prevents single-trade ruin."),
    ("• Positions",
     "Manual log of open positions. Real-time mark and P&L populated by the refresh script "
     "(if the option ticker is provided)."),
    ("• Journal",
     "Trade journal — date, ticker, thesis, outcome, lesson. Use it. The traders who "
     "survive are the ones who learn from each trade."),
    ("",  ""),
    ("REALITY CHECK FOR SMALL ACCOUNTS",
     ""),
    ("Pattern Day Trader rule",
     "Under $25k in a margin account, you're limited to 3 day-trades per 5 rolling "
     "business days. Cash accounts avoid PDT but funds settle T+1. Plan accordingly."),
    ("Cash-secured puts / covered calls",
     "Require ~$100 × strike of capital (or 100 shares for CC). At $1k account size, "
     "viable underlyings are limited to sub-$10 stocks where premiums are minimal."),
    ("Position sizing",
     "Standard guidance: never risk more than 1–5% of account on a single trade. "
     "On $1k that's $10–50 max risk per trade. The PositionSizer sheet enforces this."),
    ("",  ""),
    ("SUPPORT",
     "If a refresh fails, check the 'Last Run' log on Config. Most common issues: "
     "(1) API key missing/invalid, (2) rate limit hit (free tier = 5 calls/min), "
     "(3) ticker has no options listed, (4) network error."),
]

r = 3
for left, right in readme_rows:
    if left == "" and right == "":
        r += 1
        continue
    is_section = right == ""
    c1 = ws.cell(row=r, column=2, value=left)
    if is_section:
        c1.font = Font(name=FONT_NAME, bold=True, size=12, color="C00000")
    else:
        c1.font = Font(name=FONT_NAME, bold=True, size=10)
    c1.alignment = LEFT
    if right:
        c2 = ws.cell(row=r, column=2, value=f"{left}  —  {right}" if not is_section else right)
        c2.font = BODY_FONT
        c2.alignment = LEFT
    ws.row_dimensions[r].height = max(20, 14 * (1 + len(right) // 90))
    r += 1


# =================================================================
# SHEET 2: Config
# =================================================================
ws = wb.create_sheet("Config")
set_col_widths(ws, [32, 40, 50])
write_title(ws, 1, 1, "Configuration", span=3)

cfg_rows = [
    ("API key (Polygon.io)",       "PASTE_YOUR_API_KEY_HERE",
     "From https://polygon.io/dashboard/api-keys"),
    ("Account size ($)",           1000,
     "Total capital allocated to this strategy."),
    ("Max risk per trade (%)",     0.05,
     "Stop-loss as % of account. 0.01–0.05 is typical."),
    ("Max risk per trade ($)",     "=B5*B6",
     "Auto-computed from account size × risk %."),
    ("Max concurrent positions",   3,
     "How many open trades at once. 3–5 typical for small accounts."),
    ("Default expiration target",  "weekly",
     "weekly | 30d | 45d | quarterly — used by refresh helper."),
    ("Refresh: include S&P 500?",  "no",
     "yes/no. 'yes' uses heavy API quota; 'no' uses Watchlist only."),
    ("",                           "",                ""),
    ("Last refresh (UTC)",         "—",
     "Auto-updated by polygon_refresh.py."),
    ("Last refresh status",        "—",
     "OK / Error message."),
    ("Tickers refreshed",          0,
     "Count of tickers updated in last run."),
    ("API calls used",             0,
     "Approximate count of Polygon API calls in last run."),
]

r = 3
ws.cell(row=r, column=1, value="Parameter").font = HEADER_FONT
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.cell(row=r, column=2, value="Value").font = HEADER_FONT
ws.cell(row=r, column=2).fill = HEADER_FILL
ws.cell(row=r, column=3, value="Notes").font = HEADER_FONT
ws.cell(row=r, column=3).fill = HEADER_FILL
for c in range(1, 4):
    ws.cell(row=r, column=c).alignment = CENTER
    ws.cell(row=r, column=c).border = BOX
r = 4
for label, value, notes in cfg_rows:
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT_NAME, bold=True, size=10)
    ws.cell(row=r, column=1).alignment = LEFT
    ws.cell(row=r, column=1).border = BOX
    cell = ws.cell(row=r, column=2, value=value)
    cell.alignment = LEFT
    cell.border = BOX
    if isinstance(value, str) and value.startswith("="):
        cell.font = FORMULA_FONT
    elif label.startswith("Last") or label in ("Tickers refreshed", "API calls used"):
        cell.font = BODY_FONT  # filled by script
    else:
        cell.font = INPUT_FONT
        cell.fill = INPUT_FILL
    ws.cell(row=r, column=3, value=notes).font = NOTE_FONT
    ws.cell(row=r, column=3).alignment = LEFT
    ws.cell(row=r, column=3).border = BOX
    r += 1

ws.cell(row=4, column=2).number_format = "@"        # API key as text
ws.cell(row=5, column=2).number_format = '$#,##0;($#,##0);-'
ws.cell(row=6, column=2).number_format = "0.0%"
ws.cell(row=7, column=2).number_format = '$#,##0;($#,##0);-'
ws.cell(row=8, column=2).number_format = "0"

# Named ranges so other sheets can reference cleanly
wb.defined_names["api_key"]      = DefinedName("api_key",      attr_text="Config!$B$4")
wb.defined_names["account_size"] = DefinedName("account_size", attr_text="Config!$B$5")
wb.defined_names["risk_pct"]     = DefinedName("risk_pct",     attr_text="Config!$B$6")
wb.defined_names["max_risk_dollars"] = DefinedName("max_risk_dollars", attr_text="Config!$B$7")


# =================================================================
# SHEET 3: Watchlist
# =================================================================
ws = wb.create_sheet("Watchlist")
widths = [10, 14, 12, 12, 10, 12, 12, 12, 12, 14, 14, 14, 12, 26]
set_col_widths(ws, widths)
write_title(ws, 1, 1, "Watchlist — Live Quotes (refreshed by polygon_refresh.py)", span=14)

headers = ["Ticker", "Last", "Change", "Change %", "Bid", "Ask",
           "Day High", "Day Low", "Volume", "Prev Close",
           "30D HV %", "IV Rank %", "Earnings", "Notes"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=i, value=h)
    style_header(c)

# Pre-populate with major tickers (most liquid options on US markets)
default_watchlist = [
    "SPY", "QQQ", "IWM", "DIA", "VIX",
    "AAPL", "MSFT", "NVDA", "AMZN", "META", "GOOGL", "TSLA",
    "AMD", "AVGO", "NFLX", "CRM", "ORCL", "ADBE",
    "JPM", "BAC", "GS", "MS",
    "XLF", "XLE", "XLK", "GLD", "TLT", "USO",
    "F", "SOFI", "PLTR", "NIO",   # cheap names viable for $1k account CSPs
]
for i, t in enumerate(default_watchlist):
    r = 4 + i
    ws.cell(row=r, column=1, value=t).font = Font(name=FONT_NAME, bold=True, size=10)
    ws.cell(row=r, column=1).alignment = CENTER
    # Other columns will be filled by refresh script — leave blank
    for c in range(2, 15):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).alignment = RIGHT
    ws.cell(row=r, column=14).alignment = LEFT  # notes
    if i % 2 == 1:
        for c in range(1, 15):
            ws.cell(row=r, column=c).fill = BAND_FILL

# Number formats for watchlist columns
N = len(default_watchlist) + 4 - 1
for r in range(4, N + 1):
    ws.cell(row=r, column=2).number_format = '$#,##0.00;($#,##0.00);-'   # Last
    ws.cell(row=r, column=3).number_format = '$#,##0.00;($#,##0.00);-'   # Change
    ws.cell(row=r, column=4).number_format = '0.00%;(0.00%);-'           # Change %
    ws.cell(row=r, column=5).number_format = '$#,##0.00;($#,##0.00);-'   # Bid
    ws.cell(row=r, column=6).number_format = '$#,##0.00;($#,##0.00);-'   # Ask
    ws.cell(row=r, column=7).number_format = '$#,##0.00;($#,##0.00);-'   # High
    ws.cell(row=r, column=8).number_format = '$#,##0.00;($#,##0.00);-'   # Low
    ws.cell(row=r, column=9).number_format = '#,##0;-#,##0;-'            # Volume
    ws.cell(row=r, column=10).number_format = '$#,##0.00;($#,##0.00);-'  # Prev Close
    ws.cell(row=r, column=11).number_format = '0.0%;(0.0%);-'            # HV
    ws.cell(row=r, column=12).number_format = '0.0%;(0.0%);-'            # IV Rank

# Conditional formatting on Change % — green positive, red negative
ws.conditional_formatting.add(f"D4:D{N}",
    CellIsRule(operator="greaterThan", formula=["0"],
               fill=PatternFill("solid", start_color="C6EFCE")))
ws.conditional_formatting.add(f"D4:D{N}",
    CellIsRule(operator="lessThan", formula=["0"],
               fill=PatternFill("solid", start_color="FFC7CE")))

ws.freeze_panes = "B4"


# =================================================================
# SHEET 4: OptionsChain
# =================================================================
ws = wb.create_sheet("OptionsChain")
set_col_widths(ws, [12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12])
write_title(ws, 1, 1, "Options Chain Viewer", span=13)

# Selector area
ws.cell(row=3, column=1, value="Ticker").font = Font(name=FONT_NAME, bold=True)
sel_ticker = ws.cell(row=3, column=2, value="SPY")
sel_ticker.font = INPUT_FONT
sel_ticker.fill = INPUT_FILL
sel_ticker.alignment = CENTER
sel_ticker.border = BOX

ws.cell(row=3, column=3, value="Expiration (YYYY-MM-DD)").font = Font(name=FONT_NAME, bold=True)
sel_exp = ws.cell(row=3, column=4, value="")
sel_exp.font = INPUT_FONT
sel_exp.fill = INPUT_FILL
sel_exp.alignment = CENTER
sel_exp.border = BOX
sel_exp.comment = Comment(
    "Leave blank for nearest weekly. Format YYYY-MM-DD for specific expiry. "
    "Re-run polygon_refresh.py after changing.",
    "Dashboard")

ws.cell(row=3, column=5, value="Underlying $").font = Font(name=FONT_NAME, bold=True)
und_px = ws.cell(row=3, column=6, value="")
und_px.font = FORMULA_FONT
und_px.alignment = CENTER
und_px.border = BOX
und_px.number_format = '$#,##0.00;($#,##0.00);-'

ws.cell(row=3, column=7, value="Last refresh").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=3, column=8, value="—").font = BODY_FONT
ws.cell(row=3, column=8).alignment = CENTER
ws.cell(row=3, column=8).border = BOX

# Chain headers — calls on left, strike center, puts on right (standard layout)
hdr_row = 5
labels = [
    ("Call Last", "C"), ("Call Bid", "C"), ("Call Ask", "C"),
    ("Call Vol", "C"), ("Call OI", "C"), ("Call IV", "C"),
    ("Strike", "K"),
    ("Put Last", "P"), ("Put Bid", "P"), ("Put Ask", "P"),
    ("Put Vol", "P"), ("Put OI", "P"), ("Put IV", "P"),
]
for i, (label, kind) in enumerate(labels, 1):
    c = ws.cell(row=hdr_row, column=i, value=label)
    if kind == "K":
        c.fill = PatternFill("solid", start_color="C00000")
    elif kind == "C":
        c.fill = PatternFill("solid", start_color="2E75B6")
    else:
        c.fill = PatternFill("solid", start_color="A6A6A6")
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BOX

# Reserve 60 strike rows for the refresh script
N_STRIKES = 60
for r in range(hdr_row + 1, hdr_row + 1 + N_STRIKES):
    for col in range(1, 14):
        cc = ws.cell(row=r, column=col)
        cc.font = BODY_FONT
        cc.alignment = CENTER
        cc.number_format = '$#,##0.00;($#,##0.00);-'
        if col == 7:
            cc.font = Font(name=FONT_NAME, bold=True, size=10)
            cc.fill = PatternFill("solid", start_color="FFF2CC")
        if col in (4, 5, 11, 12):
            cc.number_format = '#,##0;-#,##0;-'
        if col in (6, 13):
            cc.number_format = '0.0%;(0.0%);-'

ws.freeze_panes = "A6"


# =================================================================
# SHEET 5: StrategyCalc
# =================================================================
ws = wb.create_sheet("StrategyCalc")
set_col_widths(ws, [22, 14, 14, 14, 14, 14, 22, 14])
write_title(ws, 1, 1, "Strategy P&L Calculator", span=8)

# Inputs box
r = 3
ws.cell(row=r, column=1, value="Underlying ($)").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=r, column=2, value=100).font = INPUT_FONT
ws.cell(row=r, column=2).fill = INPUT_FILL
ws.cell(row=r, column=2).number_format = '$#,##0.00'
ws.cell(row=r, column=2).border = BOX

ws.cell(row=r, column=4, value="Contracts").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=r, column=5, value=1).font = INPUT_FONT
ws.cell(row=r, column=5).fill = INPUT_FILL
ws.cell(row=r, column=5).border = BOX

# Legs table
r = 5
leg_headers = ["Leg", "Type", "Action", "Strike", "Premium", "Qty (× contracts)", "Position Cost", "Position Value @ Expiry"]
for i, h in enumerate(leg_headers, 1):
    style_header(ws.cell(row=r, column=i, value=h))

leg_rows_start = 6
N_LEGS = 4
example_legs = [
    ("Leg 1", "Call", "Buy",  105, 2.50, 1),
    ("Leg 2", "Call", "Sell", 110, 1.00, 1),
    ("Leg 3", "—",    "—",    "",   "",   ""),
    ("Leg 4", "—",    "—",    "",   "",   ""),
]
for i, (lname, ltype, lact, strk, prem, qty) in enumerate(example_legs):
    r = leg_rows_start + i
    ws.cell(row=r, column=1, value=lname).font = Font(name=FONT_NAME, bold=True)
    for col, v, is_input in [(2, ltype, True), (3, lact, True), (4, strk, True),
                              (5, prem, True), (6, qty, True)]:
        cell = ws.cell(row=r, column=col, value=v)
        if is_input:
            cell.font = INPUT_FONT
            cell.fill = INPUT_FILL
        cell.border = BOX
        cell.alignment = CENTER
        if col in (4, 5):
            cell.number_format = '$#,##0.00;($#,##0.00);-'
    # Position Cost = if Buy, +premium*100*qty*contracts; if Sell, -premium*100*qty*contracts
    cost_formula = (f'=IF(C{r}="Buy",E{r}*100*F{r}*$E$3,'
                    f'IF(C{r}="Sell",-E{r}*100*F{r}*$E$3,0))')
    cell = ws.cell(row=r, column=7, value=cost_formula)
    cell.font = FORMULA_FONT
    cell.number_format = '$#,##0.00;($#,##0.00);-'
    cell.border = BOX
    # Value at expiry = intrinsic value × 100 × qty × contracts × sign(action)
    # Call intrinsic = MAX(S-K, 0); Put intrinsic = MAX(K-S, 0)
    value_formula = (
        f'=IF(F{r}="","",'
        f'IF(B{r}="Call",MAX($B$3-D{r},0),'
        f'IF(B{r}="Put",MAX(D{r}-$B$3,0),0))*100*F{r}*$E$3*'
        f'IF(C{r}="Buy",1,IF(C{r}="Sell",-1,0)))'
    )
    cell = ws.cell(row=r, column=8, value=value_formula)
    cell.font = FORMULA_FONT
    cell.number_format = '$#,##0.00;($#,##0.00);-'
    cell.border = BOX

# Summary
r = leg_rows_start + N_LEGS + 1
style_header(ws.cell(row=r, column=1, value="Summary @ Underlying Price"), fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

r += 1
ws.cell(row=r, column=1, value="Net Debit / (Credit) at entry").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=r, column=2, value=f"=SUM(G{leg_rows_start}:G{leg_rows_start+N_LEGS-1})")
ws.cell(row=r, column=2).font = FORMULA_FONT
ws.cell(row=r, column=2).number_format = '$#,##0.00;($#,##0.00);-'

r += 1
ws.cell(row=r, column=1, value="P&L at expiry @ Underlying").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=r, column=2,
        value=f"=SUM(H{leg_rows_start}:H{leg_rows_start+N_LEGS-1})-SUM(G{leg_rows_start}:G{leg_rows_start+N_LEGS-1})")
ws.cell(row=r, column=2).font = FORMULA_FONT
ws.cell(row=r, column=2).number_format = '$#,##0.00;($#,##0.00);-'

# Payoff table — shows P&L across price range
r += 3
ws.cell(row=r, column=1, value="Payoff Diagram (P&L by Underlying Price)").font = Font(name=FONT_NAME, bold=True, size=12)
r += 1
style_header(ws.cell(row=r, column=1, value="Underlying $"))
style_header(ws.cell(row=r, column=2, value="Total P&L"))
style_header(ws.cell(row=r, column=3, value="% of Account"))

payoff_start_row = r + 1
# Build 21 price points from -20% to +20% of input underlying (B3)
N_POINTS = 21
for i in range(N_POINTS):
    rr = payoff_start_row + i
    # Price = B3 * (0.8 + i*0.02)
    price_formula = f"=$B$3*(0.8+{i}*0.02)"
    ws.cell(row=rr, column=1, value=price_formula).font = FORMULA_FONT
    ws.cell(row=rr, column=1).number_format = '$#,##0.00'

    # P&L at that price = sum over legs of (intrinsic - premium paid)
    # We can't easily reuse the leg table since it's tied to B3. Inline:
    pnl_parts = []
    for li in range(N_LEGS):
        lr = leg_rows_start + li
        # Intrinsic value at this price (column A in this row)
        intrinsic = (
            f'IF(B{lr}="Call",MAX(A{rr}-D{lr},0),'
            f'IF(B{lr}="Put",MAX(D{lr}-A{rr},0),0))'
        )
        # Direction: Buy=+1, Sell=-1
        direction = f'IF(C{lr}="Buy",1,IF(C{lr}="Sell",-1,0))'
        # Per-leg P&L = (intrinsic - premium) * direction * 100 * qty * contracts
        leg_pnl = (
            f'(({intrinsic})-E{lr})*({direction})*100*'
            f'IF(F{lr}="",0,F{lr})*$E$3'
        )
        pnl_parts.append(leg_pnl)
    pnl_formula = "=" + "+".join(pnl_parts)
    ws.cell(row=rr, column=2, value=pnl_formula).font = FORMULA_FONT
    ws.cell(row=rr, column=2).number_format = '$#,##0.00;($#,##0.00);-'
    ws.cell(row=rr, column=3, value=f"=B{rr}/account_size").font = FORMULA_FONT
    ws.cell(row=rr, column=3).number_format = '0.0%;(0.0%);-'

# Conditional fmt on payoff P&L column
ws.conditional_formatting.add(f"B{payoff_start_row}:B{payoff_start_row+N_POINTS-1}",
    CellIsRule(operator="greaterThan", formula=["0"],
               fill=PatternFill("solid", start_color="C6EFCE")))
ws.conditional_formatting.add(f"B{payoff_start_row}:B{payoff_start_row+N_POINTS-1}",
    CellIsRule(operator="lessThan", formula=["0"],
               fill=PatternFill("solid", start_color="FFC7CE")))

# Strategy templates note
note_r = payoff_start_row + N_POINTS + 2
ws.cell(row=note_r, column=1,
        value="Strategy templates (overwrite legs above to use):").font = Font(name=FONT_NAME, bold=True, italic=True)
templates = [
    ("Long Call",        "Leg1: Call/Buy/strike=ATM/premium=mkt. Bullish, defined risk = premium paid."),
    ("Long Put",         "Leg1: Put/Buy/strike=ATM/premium=mkt. Bearish, defined risk = premium paid."),
    ("Bull Call Spread", "Leg1: Call/Buy/lower-K. Leg2: Call/Sell/higher-K. Defined risk + capped reward."),
    ("Bear Put Spread",  "Leg1: Put/Buy/higher-K. Leg2: Put/Sell/lower-K. Defined risk + capped reward."),
    ("Iron Condor",      "Sell OTM put, Buy further-OTM put, Sell OTM call, Buy further-OTM call. Range-bound."),
    ("Cash-Secured Put", "Leg1: Put/Sell/strike below price. Collateral = strike×100. Need ~$strike×100 cash."),
    ("Covered Call",     "Need 100 shares of underlying first. Leg1: Call/Sell/above cost basis."),
    ("Long Straddle",    "Leg1: Call/Buy/ATM. Leg2: Put/Buy/ATM. Bets on big move either direction."),
]
for i, (name, desc) in enumerate(templates):
    rr = note_r + 1 + i
    ws.cell(row=rr, column=1, value=name).font = Font(name=FONT_NAME, bold=True, size=9)
    ws.cell(row=rr, column=2, value=desc).font = NOTE_FONT
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=8)
    ws.cell(row=rr, column=2).alignment = LEFT


# =================================================================
# SHEET 6: PositionSizer
# =================================================================
ws = wb.create_sheet("PositionSizer")
set_col_widths(ws, [28, 18, 18, 18, 18, 28])
write_title(ws, 1, 1, "Position Sizer — Don't Blow Up Your Account", span=6)

ws.cell(row=3, column=1, value="Account size ($)").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=3, column=2, value="=account_size").font = LINK_FONT
ws.cell(row=3, column=2).number_format = '$#,##0;($#,##0);-'

ws.cell(row=4, column=1, value="Max risk per trade (%)").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=4, column=2, value="=risk_pct").font = LINK_FONT
ws.cell(row=4, column=2).number_format = '0.0%'

ws.cell(row=5, column=1, value="Max risk per trade ($)").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=5, column=2, value="=B3*B4").font = FORMULA_FONT
ws.cell(row=5, column=2).number_format = '$#,##0.00;($#,##0.00);-'

# Sizing table — for various option premium prices
r = 7
style_header(ws.cell(row=r, column=1, value="Option Premium ($)"))
style_header(ws.cell(row=r, column=2, value="Cost per Contract ($)"))
style_header(ws.cell(row=r, column=3, value="Max Contracts (whole)"))
style_header(ws.cell(row=r, column=4, value="Total Capital Used ($)"))
style_header(ws.cell(row=r, column=5, value="% of Account"))
style_header(ws.cell(row=r, column=6, value="Notes"))

premiums = [0.05, 0.10, 0.25, 0.50, 0.75, 1.00, 1.50, 2.00, 3.00, 5.00, 7.50, 10.00]
for i, p in enumerate(premiums):
    rr = 8 + i
    ws.cell(row=rr, column=1, value=p).number_format = '$#,##0.00'
    ws.cell(row=rr, column=2, value=f"=A{rr}*100").number_format = '$#,##0.00'
    # Contracts = floor(max_risk / cost_per_contract)
    ws.cell(row=rr, column=3, value=f"=IFERROR(INT($B$5/B{rr}),0)").number_format = '#,##0;-#,##0;-'
    ws.cell(row=rr, column=4, value=f"=B{rr}*C{rr}").number_format = '$#,##0.00'
    ws.cell(row=rr, column=5, value=f"=IFERROR(D{rr}/$B$3,0)").number_format = '0.0%'
    note = ""
    if p <= 0.10:
        note = "Cheap weekly OTM — common 'lottery' premium"
    elif p <= 0.50:
        note = "Typical for OTM weekly calls/puts on low-priced names"
    elif p <= 2.00:
        note = "Near-the-money short-dated"
    else:
        note = "ITM or longer-dated — consumes more buying power"
    ws.cell(row=rr, column=6, value=note).font = NOTE_FONT
    if i % 2 == 1:
        for col in range(1, 7):
            ws.cell(row=rr, column=col).fill = BAND_FILL

# Warning box
warn_r = 8 + len(premiums) + 2
ws.cell(row=warn_r, column=1,
        value="⚠ Position Sizing Discipline").font = Font(name=FONT_NAME, bold=True, color="C00000", size=12)
ws.merge_cells(start_row=warn_r, start_column=1, end_row=warn_r, end_column=6)

warnings = [
    "On a $1k account with 5% max risk per trade ($50), you can buy 1 contract at ≤$0.50 premium without exceeding sizing rules.",
    "If a trade idea requires more contracts than the table allows, the trade is too big for your account — pass on it.",
    "The 1–5% rule exists because ~50% of trades lose money even for skilled traders. With 2% risk per trade, it takes 50 consecutive losers to wipe out — survivable. With 50% risk per trade, it takes 2.",
    "Resist the urge to 'go bigger to get back to even' after a loss. That's the mathematical path to zero.",
]
for i, w in enumerate(warnings):
    rr = warn_r + 1 + i
    ws.cell(row=rr, column=1, value="• " + w).font = BODY_FONT
    ws.cell(row=rr, column=1).alignment = LEFT
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
    ws.row_dimensions[rr].height = 30


# =================================================================
# SHEET 7: Positions
# =================================================================
ws = wb.create_sheet("Positions")
set_col_widths(ws, [12, 14, 22, 12, 12, 14, 14, 14, 14, 14, 14, 12, 30])
write_title(ws, 1, 1, "Open Positions Tracker", span=13)

pos_headers = ["Open Date", "Underlying", "Option Ticker", "Type", "Action",
               "Strike", "Expiry", "Entry Px", "Mark", "Qty",
               "Cost Basis", "Current P&L", "Notes"]
for i, h in enumerate(pos_headers, 1):
    style_header(ws.cell(row=3, column=i, value=h))

# Pre-create 15 empty rows with formulas
for i in range(15):
    rr = 4 + i
    # Cost basis = entry_px * 100 * qty * sign(action)
    ws.cell(row=rr, column=11,
            value=f'=IFERROR(IF(E{rr}="Buy",H{rr}*100*J{rr},'
                  f'IF(E{rr}="Sell",-H{rr}*100*J{rr},0)),0)')
    ws.cell(row=rr, column=11).number_format = '$#,##0.00;($#,##0.00);-'
    ws.cell(row=rr, column=11).font = FORMULA_FONT
    # P&L = (mark - entry) * 100 * qty * sign(action)
    ws.cell(row=rr, column=12,
            value=f'=IFERROR(IF(E{rr}="Buy",(I{rr}-H{rr})*100*J{rr},'
                  f'IF(E{rr}="Sell",(H{rr}-I{rr})*100*J{rr},0)),0)')
    ws.cell(row=rr, column=12).number_format = '$#,##0.00;($#,##0.00);-'
    ws.cell(row=rr, column=12).font = FORMULA_FONT
    # Format input columns
    for col in [1, 2, 3, 4, 5, 6, 7, 8, 10, 13]:
        ws.cell(row=rr, column=col).font = INPUT_FONT
        ws.cell(row=rr, column=col).fill = INPUT_FILL
        ws.cell(row=rr, column=col).border = BOX
    # Mark column populated by refresh script
    ws.cell(row=rr, column=9).font = LINK_FONT
    ws.cell(row=rr, column=9).border = BOX
    ws.cell(row=rr, column=9).number_format = '$#,##0.00;($#,##0.00);-'
    ws.cell(row=rr, column=8).number_format = '$#,##0.00;($#,##0.00);-'
    ws.cell(row=rr, column=6).number_format = '$#,##0.00;($#,##0.00);-'
    ws.cell(row=rr, column=7).number_format = "yyyy-mm-dd"
    ws.cell(row=rr, column=1).number_format = "yyyy-mm-dd"

# Totals row
tot_r = 4 + 15 + 1
ws.cell(row=tot_r, column=10, value="TOTAL").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=tot_r, column=11, value=f"=SUM(K4:K{4+14})").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=tot_r, column=11).number_format = '$#,##0.00;($#,##0.00);-'
ws.cell(row=tot_r, column=12, value=f"=SUM(L4:L{4+14})").font = Font(name=FONT_NAME, bold=True)
ws.cell(row=tot_r, column=12).number_format = '$#,##0.00;($#,##0.00);-'

ws.freeze_panes = "B4"


# =================================================================
# SHEET 8: Journal
# =================================================================
ws = wb.create_sheet("Journal")
set_col_widths(ws, [12, 12, 14, 30, 12, 14, 30, 30])
write_title(ws, 1, 1, "Trade Journal — log every trade, reread monthly", span=8)

j_headers = ["Date", "Ticker", "Strategy", "Thesis (why?)",
             "Outcome", "P&L ($)", "What I learned", "Mistake to fix"]
for i, h in enumerate(j_headers, 1):
    style_header(ws.cell(row=3, column=i, value=h))
for r in range(4, 30):
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = BOX
        ws.cell(row=r, column=c).font = INPUT_FONT
        ws.cell(row=r, column=c).fill = INPUT_FILL
        ws.cell(row=r, column=c).alignment = LEFT
    ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=6).number_format = '$#,##0.00;($#,##0.00);-'
    ws.row_dimensions[r].height = 24

ws.freeze_panes = "B4"

# =================================================================
# Save
# =================================================================
import os, sys
out_path = sys.argv[1] if len(sys.argv) > 1 else "options_dashboard.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
